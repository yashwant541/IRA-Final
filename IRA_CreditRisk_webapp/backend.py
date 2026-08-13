"""
IRA Credit Risk - Dataiku web app backend  (v2)
===============================================
Flow:  user details + file upload  ->  data checks  ->  results & override  ->
       save output to a managed folder, trigger a Tableau scenario, approve.

- The IRA engine (lib/python/IRA) is imported LAZILY and never modified.
- Uploads are in-browser (temporary).  OUTPUTS + overrides + run history are
  written to ONE Dataiku managed folder (configurable below), under subpaths:
      outputs/<run_id>.xlsx     the formatted workbook
      overrides/<run_id>.csv    the per Product+Country override table (point 6)
      runs/<run_id>.json        run metadata (user, quarter, year, time, status)
- A button triggers a Dataiku scenario (e.g. push to Tableau).
"""
import io, os, csv, json, base64, datetime, traceback
from typing import Dict, List, Any, Optional
import pandas as pd

# ------------------------------------------------------------------ CONFIG ---
STORE_FOLDER = "IRA_STORE"                 # managed folder (name or id)
TABLEAU_SCENARIO = "UPLOAD_TO_TABLEAU"     # scenario id for the Tableau push
PROJECT_KEY = None                         # None = current/default project
# -----------------------------------------------------------------------------

CATEGORIES = ["Secured", "Unsecured", "SME Banking", "Wealth Lending"]
RATING_ORDER = ["Very Low", "Low", "Medium", "High", "Very High", "Not Available"]
FINAL_CALC_LABEL = "Calculated Inherent Credit Risk Assessment:"
FINAL_OVERRIDE_LABEL = "Final Inherent Credit Risk Assessment (with Override):"
RUN_CACHE: Dict[str, Dict[str, Any]] = {}

_IRA_CACHE: Dict[str, Any] = {}
def _ira():
    if not _IRA_CACHE:
        try:
            from IRA import (ira_loaders as L, ira_build as B,
                             ira_intermediate as I, ira_detect as DET)
        except Exception:
            import ira_loaders as L, ira_build as B, ira_intermediate as I, ira_detect as DET
        _IRA_CACHE.update(L=L, B=B, I=I, DET=DET)
    return _IRA_CACHE

# ------------------------------------------------------------ store helpers ---
class _DkuStore:
    def __init__(self, folder): self.f = folder
    def put(self, path, data): self.f.upload_data(path, data)
    def get(self, path):
        try:
            with self.f.get_download_stream(path) as s: return s.read()
        except Exception: return None
    def list(self, prefix):
        try: return [p for p in self.f.list_paths_in_partition() if p.lstrip("/").startswith(prefix)]
        except Exception: return []

class _LocalStore:
    def __init__(self, root): self.root = root; os.makedirs(root, exist_ok=True)
    def put(self, path, data):
        full = os.path.join(self.root, path); os.makedirs(os.path.dirname(full), exist_ok=True)
        open(full, "wb").write(data)
    def get(self, path):
        full = os.path.join(self.root, path)
        return open(full, "rb").read() if os.path.exists(full) else None
    def list(self, prefix):
        base = os.path.join(self.root, prefix); out = []
        if os.path.isdir(base):
            for n in os.listdir(base): out.append(prefix + n)
        return out

def _store():
    try:
        import dataiku
        return _DkuStore(dataiku.Folder(STORE_FOLDER))
    except Exception:
        return _LocalStore(os.environ.get("IRA_STORE_DIR", "./ira_store"))

def _trigger_scenario(scenario_id):
    try:
        import dataiku
        client = dataiku.api_client()
        project = client.get_project(PROJECT_KEY) if PROJECT_KEY else client.get_default_project()
        scen = project.get_scenario(scenario_id)
        try: scen.run_and_wait()
        except Exception: scen.run()
        return {"ok": True, "scenario": scenario_id, "message": "Scenario triggered."}
    except Exception as ex:
        return {"ok": False, "scenario": scenario_id, "error": str(ex)}

# ------------------------------------------------------------- io utilities ---
def _read_excel_sheets(raw):
    try:
        xls = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None, engine="openpyxl")
    except Exception:
        xls = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None)
    return {name: df.values.tolist() for name, df in xls.items()}

def _parse_config(raw):
    out = {}
    for row in csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
        cat = (row.get("Category") or "").strip(); country = (row.get("Country") or "").strip()
        inc = (row.get("Include") or "").strip().lower() in ("yes","y","true","1","t","x")
        if cat and country and inc: out.setdefault(cat, []).append(country)
    return out

def _cell(v):
    return None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

def _mk_run_id(user, quarter, year, now):
    safe = "".join(ch for ch in str(user) if ch.isalnum()) or "user"
    return f"{year}_{quarter}_{safe}_{now.strftime('%Y%m%d_%H%M%S')}"

# ------------------------------------------------------------- the pipeline ---
def run_analysis(mi, other, config, user, quarter, year):
    now = datetime.datetime.now()
    run_id = _mk_run_id(user, quarter, year, now)
    availability = [
        {"table": "MI file (main workbook)", "available": bool(mi)},
        {"table": "Other tables (reference)", "available": bool(other)},
        {"table": "Countries config", "available": bool(config)},
    ]
    if not mi or not config:
        return {"ok": False, "run_id": run_id, "availability": availability,
                "error": "The MI file and the countries config are both required."}

    sheets = _read_excel_sheets(mi)
    if other:
        for nm, rows in _read_excel_sheets(other).items():
            sheets[f"OtherTables::{nm}"] = rows
    cfg = _parse_config(config)

    M = _ira(); L, B = M["L"], M["B"]
    tables = L.load_tables(sheets)
    per_cat = B.resolve_countries(tables, None, cfg)
    frames = B.build_all(tables, countries_per_category=cfg)
    inter = B.build_intermediate_frames(tables, per_cat)
    mapping = B.build_mapping()

    for name, present in (
        ("Sovereign rating table", bool(tables.get("sovereign"))),
        ("Dispensations tables", bool(tables.get("dispensations"))),
        ("CRA breaches tables", bool(tables.get("cra_breaches"))),
        ("Property Price Index", tables.get("PPI") is not None),
        ("Interest rates", tables.get("interest_rates") is not None),
    ):
        availability.append({"table": name, "available": present})

    results, calculated, na_details, countries_by_product = {}, {}, [], {}
    for cat in CATEGORIES:
        df = frames.get(f"IRA - {cat}")
        results[cat], calculated[cat] = {}, {}
        countries_by_product[cat] = list(per_cat.get(cat, []))
        if df is None or df.empty: continue
        for country, grp in df.groupby("Country", sort=False):
            rows = []
            for _, r in grp.iterrows():
                label = str(r["Label"]); rating = _cell(r["Risk Rating"])
                rec = {"label": label, "value": _cell(r["Value"]),
                       "rating": rating, "number": _cell(r["Risk Number"])}
                if label.startswith("Calculated"):
                    calculated[cat][country] = {"rating": rating, "value": _cell(r["Value"])}
                else:
                    rows.append(rec)
                    if rating == "Not Available":
                        na_details.append({"product": cat, "country": country, "label": label,
                            "reason": str(r["What to do in Value Column"]).replace("Not Available - ", "")})
            calc = calculated[cat].get(country, {})
            rows.append({"label": FINAL_CALC_LABEL, "value": calc.get("value"),
                         "rating": calc.get("rating"), "number": None, "calc": True})
            rows.append({"label": FINAL_OVERRIDE_LABEL, "value": None, "rating": None,
                         "number": None, "override": True})
            results[cat][country] = rows

    meta = {"run_id": run_id, "user": user, "quarter": quarter, "year": year,
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "n_countries": len({c for cat in CATEGORIES for c in per_cat.get(cat, [])}),
            "na_count": len(na_details), "status": "processed"}
    try: excel_bytes = _excel_bytes(frames, inter, mapping)
    except Exception: excel_bytes = None

    RUN_CACHE[run_id] = {"excel": excel_bytes, "meta": meta, "results": results,
                         "calculated": calculated, "frames": frames, "inter": inter, "mapping": mapping}
    return {"ok": True, "run_id": run_id, "meta": meta, "availability": availability,
            "na_details": na_details, "countries_by_product": countries_by_product,
            "results": results, "calculated": calculated, "excel_ready": excel_bytes is not None}

# ------------------------------------------------------------- excel writer ---
def _excel_bytes(frames, inter, mapping, overrides=None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    fills = {"Very Low":"C6EFCE","Low":"D9EAD3","Medium":"FFF2CC","High":"FCE5CD",
             "Very High":"F4CCCC","Not Available":"EFEFEF"}
    ov_map = {(o["product"], o["country"]): o for o in (overrides or [])}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for cat in CATEGORIES:
            df = frames.get(f"IRA - {cat}")
            if df is None: continue
            if overrides: df = _inject_override_rows(df, cat, ov_map)
            df.to_excel(xw, sheet_name=cat[:31], index=False)
        if overrides: pd.DataFrame(overrides).to_excel(xw, sheet_name="Overrides", index=False)
        mapping.to_excel(xw, sheet_name="Mapping", index=False)
        for title, d in inter.items(): d.to_excel(xw, sheet_name=title[:31], index=False)
    buf.seek(0); wb = load_workbook(buf)
    hf = PatternFill("solid", fgColor="E8EEF7"); hfont = Font(bold=True, color="1F3A5F", size=10)
    thin = Side(style="thin", color="DCE3EF"); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ws in wb.worksheets:
        rc = None
        for c in range(1, ws.max_column+1):
            ws.cell(1,c).fill=hf; ws.cell(1,c).font=hfont
            ws.cell(1,c).alignment=Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width=22
            if str(ws.cell(1,c).value).strip()=="Risk Rating": rc=c
        ws.freeze_panes="A2"; ws.row_dimensions[1].height=26
        for r in range(2, ws.max_row+1):
            if rc:
                v=ws.cell(r,rc).value
                if v in fills:
                    for c in range(1, ws.max_column+1):
                        ws.cell(r,c).fill=PatternFill("solid", fgColor=fills[v])
            for c in range(1, ws.max_column+1): ws.cell(r,c).border=bd
    out=io.BytesIO(); wb.save(out)
    return base64.b64encode(out.getvalue()).decode("ascii")

def _inject_override_rows(df, cat, ov_map):
    rows=[]
    for country, grp in df.groupby("Country", sort=False):
        for _, r in grp.iterrows(): rows.append(r.to_dict())
        o=ov_map.get((cat, country))
        if o:
            rows.append({"Country":country, "Label":FINAL_OVERRIDE_LABEL,
                         "Value":o.get("override_text",""), "Risk Rating":o.get("override_rating") or "",
                         "Risk Number":"", "What to do in Value Column":""})
    return pd.DataFrame(rows, columns=df.columns)

# --------------------------------------------------------------- endpoints ---
try:
    from flask import request

    def _json(obj, code=200):
        return app.response_class(json.dumps(obj, default=str), mimetype="application/json", status=code)

    @app.route("/health")
    def health():
        ok=True
        try: _store()
        except Exception: ok=False
        return _json({"ok":True, "app":"IRA Credit Risk", "store":ok})

    @app.route("/analyze", methods=["POST"])
    def analyze():
        try:
            def _b(k):
                f=request.files.get(k); return f.read() if f else None
            res=run_analysis(_b("mi"), _b("other"), _b("config"),
                             request.form.get("user","").strip(),
                             request.form.get("quarter","").strip(),
                             request.form.get("year","").strip())
            return _json(res)
        except Exception as ex:
            return _json({"ok":False, "error":f"{type(ex).__name__}: {ex}", "trace":traceback.format_exc()}, 500)

    @app.route("/save_output", methods=["POST"])
    def save_output():
        b=request.get_json(force=True, silent=True) or {}; rid=b.get("run_id")
        c=RUN_CACHE.get(rid)
        if not c or not c.get("excel"): return _json({"ok":False,"error":"Run not found / output not ready."},404)
        try:
            s=_store(); s.put(f"outputs/{rid}.xlsx", base64.b64decode(c["excel"]))
            m=dict(c["meta"]); m["status"]="saved_to_folder"
            s.put(f"runs/{rid}.json", json.dumps(m).encode()); c["meta"]=m
            return _json({"ok":True, "path":f"outputs/{rid}.xlsx"})
        except Exception as ex: return _json({"ok":False,"error":str(ex)},500)

    @app.route("/approve", methods=["POST"])
    def approve():
        b=request.get_json(force=True, silent=True) or {}; rid=b.get("run_id"); ov=b.get("overrides",[])
        c=RUN_CACHE.get(rid)
        if not c: return _json({"ok":False,"error":"Run not found."},404)
        try:
            s=_store(); rows=[]
            for o in ov:
                cc=c["calculated"].get(o["product"],{}).get(o["country"],{})
                rows.append({"run_id":rid,"user":c["meta"]["user"],"quarter":c["meta"]["quarter"],
                    "year":c["meta"]["year"],"product":o["product"],"country":o["country"],
                    "calculated_rating":cc.get("rating"),"calculated_score":cc.get("value"),
                    "override_rating":o.get("override_rating") or "","override_text":o.get("override_text") or "",
                    "timestamp":c["meta"]["timestamp"]})
            buf=io.StringIO()
            w=csv.DictWriter(buf, fieldnames=(list(rows[0].keys()) if rows else
                ["run_id","product","country","override_rating","override_text"]))
            w.writeheader(); [w.writerow(r) for r in rows]
            s.put(f"overrides/{rid}.csv", buf.getvalue().encode())
            # rebuild workbook WITH override rows and save
            try:
                xls=_excel_bytes(c["frames"], c["inter"], c["mapping"], overrides=rows)
                c["excel"]=xls
            except Exception: xls=c["excel"]
            s.put(f"outputs/{rid}.xlsx", base64.b64decode(c["excel"]))
            m=dict(c["meta"]); m["status"]="approved"
            m["approved_at"]=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); m["overrides"]=len(rows)
            s.put(f"runs/{rid}.json", json.dumps(m).encode()); c["meta"]=m
            return _json({"ok":True, "overrides_saved":len(rows), "path":f"overrides/{rid}.csv"})
        except Exception as ex: return _json({"ok":False,"error":str(ex)},500)

    @app.route("/trigger_tableau", methods=["POST"])
    def trigger_tableau():
        return _json(_trigger_scenario(TABLEAU_SCENARIO))

    @app.route("/download/<run_id>")
    def download(run_id):
        c=RUN_CACHE.get(run_id); data=None
        if c and c.get("excel"): data=base64.b64decode(c["excel"])
        else: data=_store().get(f"outputs/{run_id}.xlsx")
        if not data: return _json({"ok":False,"error":"Output not found."},404)
        return app.response_class(data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="IRA_{run_id}.xlsx"'})

    @app.route("/history")
    def history():
        try:
            s=_store(); runs=[]
            for p in s.list("runs/"):
                raw=s.get(p.lstrip("/"))
                if raw:
                    try: runs.append(json.loads(raw.decode()))
                    except Exception: pass
            runs.sort(key=lambda m: m.get("timestamp",""), reverse=True)
            return _json({"ok":True, "runs":runs[:25]})
        except Exception as ex:
            return _json({"ok":False, "error":str(ex), "runs":[]})
except NameError:
    pass
